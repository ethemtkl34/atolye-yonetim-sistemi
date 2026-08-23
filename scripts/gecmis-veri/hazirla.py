#!/usr/bin/env python3
"""
Geçmiş veri aktarımı — 1. aşama: kaynakları oku, temizle, eşleştir.

Bu betik VERİTABANINA DOKUNMAZ. Excel kütüğü ile PDF rapor klasörünü okuyup
iki çıktı üretir:

    cikti/gecmis-veri.json   -> 2. aşamanın (aktar.ts) girdisi
    cikti/denetim.md         -> insan gözüyle okunacak denetim raporu
    cikti/raporsuz-ogrenciler.xlsx -> raporu bulunamayan öğrenciler

Araya JSON konmasının sebebi: aktarım canlı veritabanına yazıyor ve neyin
yazılacağı çalıştırmadan ÖNCE okunabilmeli. Denetim raporu her düzeltmeyi
(atılan telefon, eşlenen ad, velisiz öğrenci) tek tek sayar.

Çalıştırma:
    python3 scripts/gecmis-veri/hazirla.py

Bağımlılıklar: openpyxl (kurulu), pdftotext (poppler, kurulu).
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

# ---------------------------------------------------------------------------
# Ayarlar
# ---------------------------------------------------------------------------

EXCEL = Path.home() / "Downloads" / "atölye eski öğrenci dataları.xlsx"
PDF_KOK = Path.home() / "Downloads" / "sendgb-UOsCRqp6nQS"
CIKTI = Path(__file__).parent / "cikti"

# Şube eşlemesi — TEK DEĞİŞTİRME NOKTASI.
#
# Karar: bütün geçmiş veri Ümraniye'ye yazılıyor (PDF kapaklarının hepsinde
# "ÜMRANİYE - DAHİ PARK" yazıyor). Excel ana sayfası satır 229'dan sonra
# "1.Kur - Ümraniye" etiketli ayrı bir bloğa geçiyor; bu blok ayrı bir şube
# olsaydı `ALT_BLOK_SUBE` "gunesli" yapılır ve aktarım yeniden koşulurdu.
UST_BLOK_SUBE = "umraniye"
ALT_BLOK_SUBE = "umraniye"

# Dönem adları — `Term.name` olarak birebir yazılır.
D_KIS = "2025-2026 Kış 1. Kur"
D_BAHAR = "2025-2026 Bahar 2. Kur"
D_YAZ25 = "2025 Yaz Atölyesi"
D_YAZ26 = "2026 Yaz Atölyesi"
K_DRAMA = "Drama Kulübü 2026"
K_ROBOTIK = "Robotik Kodlama Kulübü 2026"

DONEM_EGITIM_YILI = {
    D_KIS: "2025-2026",
    D_BAHAR: "2025-2026",
    D_YAZ25: "2024-2025",
    D_YAZ26: "2025-2026",
}

# PDF klasörü -> (dönem adı, zaman dilimi ipucu, rapor tarihi)
# Rapor tarihi dönemin bitişine yakın bir gün; PDF'lerin kendi üretim tarihi
# tek tek okunuyor, bu yalnızca okunamazsa devreye giren yedek.
KLASOR_DONEMI = [
    ("KUR ÖĞLEDEN ÖNCE", D_KIS, "OGLEDEN_ONCE", date(2026, 1, 31)),
    ("KUR ÖĞLEDEN SONRA", D_KIS, "OGLEDEN_SONRA", date(2026, 1, 31)),
    ("cumartesi öğrenci raporlar", D_BAHAR, None, date(2026, 6, 6)),
    ("önce öğrenci raporlar", D_BAHAR, "OGLEDEN_ONCE", date(2026, 6, 6)),
    ("sonra öğrenci raporlar", D_BAHAR, "OGLEDEN_SONRA", date(2026, 6, 6)),
    ("YAZ ÖĞRENCİ RAPORLARI", D_YAZ26, None, date(2026, 8, 14)),
]

# Aynı çocuk kütüğe iki farklı yazımla girilmiş ya da PDF kapağındaki ad
# Excel'den farklı — elle doğrulanmış eşleme. Anahtar da değer de normalize
# edilmiş hâlidir; eşleşen satırlar tek öğrencide toplanır.
#
# Kütük içi mükerrerler için ölçüt: doğum tarihi VE her iki velinin telefonu
# birebir aynı olacak. Yalnız ad benzerliği yetmez — "Ahya Naz Örnek" ile
# "Meva Duru Örnek" aynı doğum tarihini ve aynı telefonları paylaşıyor ama
# ikizler, ikisinin de kendi raporu var.
#
# BİLEREK DIŞARIDA: "muhammet akif aktas" (r243) ile "muhammed akif aktas"
# (r57) da bu ölçütü karşılıyor, ama iki satır AYNI dönemlerde FARKLI
# sınıflarda (B-1 / B-2). Tek çocuk mu, iki ayrı kayıt mı olduğu kütükten
# çözülemedi; kullanıcı karar verene kadar iki öğrenci olarak duruyorlar.
AD_ESLEMESI = {
    "ahmet hasan kopucuoglu": "ahmet hasan kapucuoglu",
    "atlas balkan asdemir": "atlas balkan astemir",
    "can begde": "can bedge",
    "kivanc yoruk": "kivanc yuruk",
    "oguz kaan yalcin": "oguz kagan yalcin",
    "alya naz ornek": "ahya naz ornek",
    "ayse melis beyazit": "melis beyazit",
    # Yaz sayfasında göbek adı yazılmamış; doğum tarihi ve iki veli telefonu
    # da ana sayfadaki satırla birebir aynı.
    "alp dilsiz": "alp asaf dilsiz",
}

# Aynı ad, iki farklı doğum tarihi — elle verilmiş karar.
#   True  -> tek çocuk, tarihlerden biri hatalı (ad tek başına anahtar)
#   False -> gerçekten iki ayrı çocuk (ad + doğum tarihi birlikte anahtar)
AD_TEK_KISI = {
    "ahmet yusuf pinar": True,
    "mahir yilmaz": False,
}

# ---------------------------------------------------------------------------
# Normalizasyon — src/lib/turkce.ts ile BİREBİR aynı davranmalı
# ---------------------------------------------------------------------------

_ARAMA_KARSILIKLARI = {
    "ç": "c", "ğ": "g", "ı": "i", "ö": "o", "ş": "s", "ü": "u",
    "â": "a", "î": "i", "û": "u",
}
_TR_KUCUK = {"I": "ı", "İ": "i"}


def _tr_kucult(metin: str) -> str:
    """toLocaleLowerCase("tr-TR") karşılığı."""
    return "".join(_TR_KUCUK.get(h, h) for h in metin).lower()


def normalize_arama(metin: str) -> str:
    """`normalizeArama()` (src/lib/turkce.ts) — Student.searchName üretir."""
    metin = unicodedata.normalize("NFC", metin)
    kucuk = _tr_kucult(metin)
    sade = "".join(_ARAMA_KARSILIKLARI.get(h, h) for h in kucuk)
    return re.sub(r"\s+", " ", sade).strip()


def normalize_telefon(telefon: str) -> str:
    """`normalizeTelefon()` (src/lib/turkce.ts) — Guardian.searchPhone üretir."""
    rakamlar = re.sub(r"\D", "", telefon)
    if rakamlar.startswith("00"):
        rakamlar = rakamlar[2:]
    if rakamlar.startswith("90") and len(rakamlar) > 10:
        rakamlar = rakamlar[2:]
    if rakamlar.startswith("0"):
        rakamlar = rakamlar[1:]
    return rakamlar


# ---------------------------------------------------------------------------
# Hücre temizleyicileri
# ---------------------------------------------------------------------------

uyarilar: list[str] = []


def uyar(kategori: str, ileti: str) -> None:
    uyarilar.append(f"{kategori}\t{ileti}")


def metin(deger) -> str:
    if deger is None:
        return ""
    return unicodedata.normalize("NFC", str(deger)).strip()


def ad_soyad_ayir(tam: str) -> tuple[str, str]:
    """Son kelime soyad, kalanı ad. Tek kelimeyse soyad boş kalmaz."""
    parcalar = [p for p in re.split(r"\s+", tam.strip()) if p]
    if not parcalar:
        return ("", "")
    if len(parcalar) == 1:
        return (parcalar[0], "-")
    return (" ".join(parcalar[:-1]), parcalar[-1])


def buyuk_harfe(tam: str) -> str:
    """Excel'de karışık yazım var; hepsi tek biçime çekilir."""
    return " ".join(
        p[:1].upper() + _tr_kucult(p[1:]).upper() if False else _baslik(p)
        for p in re.split(r"\s+", tam.strip())
        if p
    )


def _baslik(kelime: str) -> str:
    if not kelime:
        return kelime
    ilk = kelime[0]
    ilk = {"i": "İ", "ı": "I"}.get(ilk, ilk.upper())
    kalan = _tr_kucult(kelime[1:])
    return ilk + kalan


def dogum_tarihi(deger, kaynak: str) -> str | None:
    """Excel seri numarası ya da gerçek tarih -> ISO. Bozuksa None."""
    if deger is None:
        return None
    if isinstance(deger, datetime):
        return deger.date().isoformat()
    if isinstance(deger, date):
        return deger.isoformat()
    ham = metin(deger)
    if not ham:
        return None
    if re.fullmatch(r"\d+(\.0)?", ham):
        seri = int(float(ham))
        # 1900 tabanı + Excel'in 1900 artık yıl hatası.
        if 20000 <= seri <= 60000:
            return (date(1899, 12, 30) + timedelta(days=seri)).isoformat()
    uyar("DOĞUM TARİHİ", f"{kaynak}: okunamayan değer {ham!r} -> boş bırakıldı")
    return None


def telefon(deger, kaynak: str) -> tuple[str | None, str | None]:
    """(phone, searchPhone). Geçersizse (None, None) + denetim satırı."""
    ham = metin(deger)
    if not ham:
        return (None, None)
    anahtar = normalize_telefon(ham)
    if len(anahtar) == 10 and anahtar.startswith("5"):
        return ("0" + anahtar, anahtar)
    uyar("TELEFON", f"{kaynak}: geçersiz numara {ham!r} -> boş bırakıldı")
    return (None, None)


SINIF_DUZELTME = re.compile(r"^(\d+)\s*\.?\s*sinif$")


def okul_sinifi(deger, kaynak: str) -> str | None:
    """'5 YAŞ' / '1.SINIF' / '1. Sınıf' karmaşasını tek biçime çeker."""
    ham = metin(deger)
    if not ham:
        return None
    anahtar = normalize_arama(ham)
    yas = re.fullmatch(r"(\d+)\s*yas", anahtar)
    if yas:
        return f"{yas.group(1)} yaş"
    sinif = SINIF_DUZELTME.match(anahtar)
    if sinif:
        return f"{sinif.group(1)}. sınıf"
    uyar("SINIF", f"{kaynak}: tanınmayan değer {ham!r} -> boş bırakıldı")
    return None


def zaman_dilimi(deger) -> str:
    anahtar = normalize_arama(metin(deger))
    if "sonra" in anahtar:
        return "OGLEDEN_SONRA"
    return "OGLEDEN_ONCE"


# ---------------------------------------------------------------------------
# Öğrenci defteri
# ---------------------------------------------------------------------------


class Defter:
    """Ad + doğum tarihine göre tekilleştirilmiş öğrenci listesi."""

    def __init__(self) -> None:
        self.ogrenciler: dict[str, dict] = {}
        self.ada_gore: dict[str, list[str]] = defaultdict(list)

    def anahtar(self, aramaAdi: str, dt: str | None) -> str:
        if AD_TEK_KISI.get(aramaAdi) is True:
            return aramaAdi
        return f"{aramaAdi}|{dt or ''}"

    def ekle(self, *, tamAd: str, dt: str | None, sube: str,
             okul: str | None, veliler: list[dict], kaynak: str) -> str:
        aramaAdi = normalize_arama(tamAd)
        aramaAdi = AD_ESLEMESI.get(aramaAdi, aramaAdi)
        anahtar = self.anahtar(aramaAdi, dt)

        mevcut = self.ogrenciler.get(anahtar)
        if mevcut is None:
            ad, soyad = ad_soyad_ayir(buyuk_harfe(tamAd))
            mevcut = {
                "anahtar": anahtar,
                "searchName": aramaAdi,
                "firstName": ad,
                "lastName": soyad,
                "birthDate": dt,
                "school": None,
                "grade": okul,
                "branchCode": sube,
                "guardians": [],
                "kaynaklar": [],
            }
            self.ogrenciler[anahtar] = mevcut
            self.ada_gore[aramaAdi].append(anahtar)
        else:
            if mevcut["birthDate"] is None and dt:
                mevcut["birthDate"] = dt
            if mevcut["grade"] is None and okul:
                mevcut["grade"] = okul
            if mevcut["branchCode"] != sube:
                uyar("ŞUBE", f"{tamAd}: iki şubede birden ({mevcut['branchCode']} / {sube})"
                             f" -> {mevcut['branchCode']} korundu")

        mevcut["kaynaklar"].append(kaynak)
        for veli in veliler:
            if any(v["type"] == veli["type"] for v in mevcut["guardians"]):
                continue
            mevcut["guardians"].append(veli)
        return anahtar

    def bul(self, tamAd: str, *, donem: str | None = None,
            dilim: str | None = None, kayitDizini=None,
            kaynak: str = "") -> str | None:
        """
        Adı öğrenciye bağlar. Aynı adda birden çok çocuk varsa (gerçekten iki
        ayrı çocuk olabiliyor) raporun dönemi ve klasörünün zaman dilimi
        ayırt edici olarak kullanılır.
        """
        aramaAdi = normalize_arama(tamAd)
        aramaAdi = AD_ESLEMESI.get(aramaAdi, aramaAdi)
        adaylar = self.ada_gore.get(aramaAdi, [])
        if not adaylar:
            return None
        if len(adaylar) == 1:
            return adaylar[0]

        if donem and kayitDizini is not None:
            oDonemde = [a for a in adaylar if kayitDizini.get((a, donem))]
            if len(oDonemde) == 1:
                return oDonemde[0]
            if dilim and len(oDonemde) > 1:
                dilimeUyan = [
                    a for a in oDonemde
                    if any(k["timeSlot"] == dilim for k in kayitDizini[(a, donem)])
                ]
                if len(dilimeUyan) == 1:
                    return dilimeUyan[0]
            if oDonemde:
                adaylar = oDonemde

        uyar("EŞLEME", f"{tamAd}: aynı adda {len(adaylar)} öğrenci var, "
                       f"{kaynak or 'kayıt'} için ayırt edilemedi -> ilki seçildi")
        return adaylar[0]


def veli_yap(tur: str, adDeger, telDeger, kaynak: str) -> dict | None:
    """
    Excel'de yalnızca ilk ad var (TUĞBA, İBRAHİM). `Guardian.fullName` zorunlu
    olduğu için değer OLDUĞU GİBİ yazılır — soyad uydurulmaz.
    """
    ad = metin(adDeger)
    if not ad:
        return None
    tel, aramaTel = telefon(telDeger, f"{kaynak} / {tur.lower()}")
    return {"type": tur, "fullName": buyuk_harfe(ad), "phone": tel, "searchPhone": aramaTel}


# ---------------------------------------------------------------------------
# Excel okuma
# ---------------------------------------------------------------------------

def excel_oku(defter: Defter) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    kayitlar: list[dict] = []

    def kayit(anahtar, program, tur, grupAdi, dilim, durum="AKTIF"):
        kayitlar.append({
            "ogrenciAnahtari": anahtar,
            "programAdi": program,
            "programTuru": tur,          # "donem" | "kulup"
            "grupAdi": grupAdi,
            "timeSlot": dilim,
            "status": durum,
        })

    # --- Ana sayfa: Kış 1. Kur + Bahar 2. Kur ---------------------------------
    ws = wb["DAHİPARK ATÖLYE KAYITLARI 2025-"]
    for i, satir in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not satir[0] or not metin(satir[0]):
            continue
        kaynak = f"Ana sayfa r{i}"
        tamAd = metin(satir[0])
        kur = metin(satir[3])
        altBlok = "mraniye" in kur
        sube = ALT_BLOK_SUBE if altBlok else UST_BLOK_SUBE
        dt = dogum_tarihi(satir[1], f"{kaynak} {tamAd}")
        sinif = metin(satir[4])
        donemBilgisi = normalize_arama(metin(satir[5]))
        dilim = zaman_dilimi(satir[6])
        veliler = [v for v in (
            veli_yap("ANNE", satir[7], satir[9], f"{kaynak} {tamAd}"),
            veli_yap("BABA", satir[8], satir[10], f"{kaynak} {tamAd}"),
        ) if v]

        anahtar = defter.ekle(
            tamAd=tamAd, dt=dt, sube=sube,
            okul=okul_sinifi(satir[2], f"{kaynak} {tamAd}"),
            veliler=veliler, kaynak=kaynak,
        )

        if kur == "İPTAL":
            # Tek satır; sınıfı "BOŞ". Öğrenci açılır, kaydı iptal yazılır.
            kayit(anahtar, D_KIS, "donem", "A-1", dilim, durum="IPTAL")
            uyar("İPTAL", f"{kaynak} {tamAd}: iptal satırı, Kış dönemine iptal kaydı olarak yazıldı")
            continue

        if sinif in ("", "BOŞ"):
            uyar("GRUP", f"{kaynak} {tamAd}: atölye sınıfı boş -> kayıt atlandı")
            continue

        if altBlok:
            # Alt blokta kur sütunu tek değer; asıl ayrım DÖNEM BİLGİSİ'nde.
            # Yazım serbest ("1.DÖNEM", "1. Dönem"), noktalama ayıklanır.
            sade = re.sub(r"[^a-z0-9]", "", donemBilgisi)
            donemler = {"1donem": [D_KIS], "2donem": [D_BAHAR],
                        "yillik": [D_KIS, D_BAHAR]}.get(sade)
        else:
            donemler = {"1.KUR": [D_KIS], "2.KUR": [D_BAHAR],
                        "1.-2.KUR": [D_KIS, D_BAHAR]}.get(kur)

        if not donemler:
            uyar("DÖNEM", f"{kaynak} {tamAd}: kur={kur!r} dönem={donemBilgisi!r} "
                          f"çözülemedi -> kayıt atlandı")
            continue

        for donem in donemler:
            kayit(anahtar, donem, "donem", sinif, dilim)

    # --- Yaz sayfası: YAZ 2025 + YAZ 2026 ------------------------------------
    ws = wb["YAZ ATÖLYE KAYITLARI"]
    for i, satir in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not satir[0] or not metin(satir[0]):
            continue
        kaynak = f"Yaz sayfası r{i}"
        tamAd = metin(satir[0])
        dt = dogum_tarihi(satir[1], f"{kaynak} {tamAd}")
        sinif = metin(satir[3])
        donemHam = normalize_arama(metin(satir[4]))
        veliler = [v for v in (
            veli_yap("ANNE", satir[5], satir[7], f"{kaynak} {tamAd}"),
            veli_yap("BABA", satir[6], satir[8], f"{kaynak} {tamAd}"),
        ) if v]

        anahtar = defter.ekle(
            tamAd=tamAd, dt=dt, sube=UST_BLOK_SUBE,
            okul=okul_sinifi(satir[2], f"{kaynak} {tamAd}"),
            veliler=veliler, kaynak=kaynak,
        )

        donem = {"yaz 2025": D_YAZ25, "yaz 2026": D_YAZ26}.get(donemHam)
        if not donem:
            uyar("DÖNEM", f"{kaynak} {tamAd}: dönem {donemHam!r} çözülemedi -> kayıt atlandı")
            continue
        if not sinif:
            uyar("GRUP", f"{kaynak} {tamAd}: grup boş -> kayıt atlandı")
            continue
        # Yaz sayfasında zaman dilimi sütunu yok.
        kayit(anahtar, donem, "donem", sinif, "OGLEDEN_ONCE")

    # --- Kulüpler -------------------------------------------------------------
    for sayfa, kulup, grupSutunu in (
        ("DRAMA KULÜP 2026 DÖNEM", K_DRAMA, None),
        ("ROBOTİK KODLAMA KULÜP 2026 ", K_ROBOTIK, 4),
    ):
        ws = wb[sayfa]
        anneSutunu = 4 if grupSutunu is None else 5
        for i, satir in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not satir[0] or not metin(satir[0]):
                continue
            kaynak = f"{kulup} r{i}"
            tamAd = metin(satir[0])
            dt = dogum_tarihi(satir[1], f"{kaynak} {tamAd}")
            veliler = [v for v in (
                veli_yap("ANNE", satir[anneSutunu], satir[anneSutunu + 2], f"{kaynak} {tamAd}"),
                veli_yap("BABA", satir[anneSutunu + 1], satir[anneSutunu + 3], f"{kaynak} {tamAd}"),
            ) if v]
            anahtar = defter.ekle(
                tamAd=tamAd, dt=dt, sube=UST_BLOK_SUBE,
                okul=okul_sinifi(satir[2], f"{kaynak} {tamAd}"),
                veliler=veliler, kaynak=kaynak,
            )
            grupAdi = "1. Grup"
            if grupSutunu is not None:
                ham = normalize_arama(metin(satir[grupSutunu]))
                grupAdi = "2. Grup" if ham.startswith("2") else "1. Grup"
            kayit(anahtar, kulup, "kulup", grupAdi, "OGLEDEN_ONCE")

    return kayitlar


# ---------------------------------------------------------------------------
# PDF okuma
# ---------------------------------------------------------------------------

def pdf_kapagi(yol: Path) -> tuple[str | None, str | None]:
    try:
        cikti = subprocess.run(
            ["pdftotext", "-f", "1", "-l", "1", "-layout", str(yol), "-"],
            capture_output=True, text=True, timeout=60,
        ).stdout
    except Exception as hata:  # noqa: BLE001
        uyar("PDF", f"{yol.name}: okunamadı ({hata})")
        return (None, None)
    ad = re.search(r"Öğrenci Adı:\s*(.+)", cikti)
    sinif = re.search(r"Sınıfı:\s*(.+)", cikti)
    return (ad.group(1).strip() if ad else None,
            sinif.group(1).strip() if sinif else None)


def pdf_tarihi(yol: Path, yedek: date) -> str:
    try:
        cikti = subprocess.run(["pdfinfo", str(yol)], capture_output=True,
                               text=True, timeout=30).stdout
        eslesme = re.search(r"CreationDate:\s*(.+)", cikti)
        if eslesme:
            return datetime.strptime(
                eslesme.group(1).strip()[:24], "%a %b %d %H:%M:%S %Y"
            ).date().isoformat()
    except Exception:  # noqa: BLE001
        pass
    return yedek.isoformat()


def klasor_donemi(bagilYol: str) -> tuple[str, str | None, date] | None:
    normal = unicodedata.normalize("NFC", bagilYol)
    for parca, donem, dilim, tarih in KLASOR_DONEMI:
        if parca in normal:
            return (donem, dilim, tarih)
    return None


def pdf_oku(defter: Defter, kayitlar: list[dict]) -> list[dict]:
    """
    Her PDF'i bir öğrenciye ve mümkünse o dönemdeki kaydına bağlar.

    Eşleştirme dosya adına DEĞİL, kapaktaki `Öğrenci Adı:` satırına dayanır —
    dosya adları çok tutarsız ("B2 Ö.Ö. MUSAB TAHA TUNÇxlsx.pdf"). Klasörün
    zaman dilimi ipucu iki yerde kullanılır: aynı adda iki öğrenci varsa
    hangisi olduğunu ayırmak için ve Excel'de karşılığı olmayan bir kayıt
    açmak gerektiğinde.
    """
    kayitDizini: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for kayit in kayitlar:
        kayitDizini[(kayit["ogrenciAnahtari"], kayit["programAdi"])].append(kayit)

    raporlar: list[dict] = []
    gorulen: set[str] = set()

    for yol in sorted(PDF_KOK.rglob("*.pdf")):
        bagil = unicodedata.normalize("NFC", str(yol.relative_to(PDF_KOK)))
        if bagil in gorulen:
            continue
        gorulen.add(bagil)

        bilgi = klasor_donemi(bagil)
        if bilgi is None:
            uyar("PDF", f"{bagil}: klasöründen dönem çözülemedi -> atlandı")
            continue
        donem, klasorDilimi, yedekTarih = bilgi

        ad, sinif = pdf_kapagi(yol)
        if not ad:
            uyar("PDF", f"{bagil}: kapakta 'Öğrenci Adı' bulunamadı -> atlandı")
            continue

        anahtar = defter.bul(ad, donem=donem, dilim=klasorDilimi,
                             kayitDizini=kayitDizini, kaynak=bagil)
        if anahtar is None:
            # Excel'de karşılığı yok: belge kaybolmasın diye öğrenci PDF
            # kapağından açılır, veli bilgisi boş kalır.
            anahtar = defter.ekle(
                tamAd=ad, dt=None, sube=UST_BLOK_SUBE, okul=None,
                veliler=[], kaynak=f"YALNIZ PDF: {bagil}",
            )
            defter.ogrenciler[anahtar]["yalnizPdf"] = True
            uyar("VELİSİZ", f"{ad}: Excel'de yok, {bagil} kapağından oluşturuldu")

        adaylar = kayitDizini[(anahtar, donem)]
        if not adaylar:
            # Rapor var ama Excel'de o döneme kayıt yok — öğrenci o dönemde
            # atölyeye GİTMİŞ, kütükte eksik kalmış. Kaydı rapordan türetiyoruz;
            # aksi hâlde belge dönemsiz kalır ve geçmiş yarım görünür.
            kayit = {
                "ogrenciAnahtari": anahtar,
                "programAdi": donem,
                "programTuru": "donem",
                "grupAdi": sinif or "A-1",
                "timeSlot": klasorDilimi or "OGLEDEN_ONCE",
                "status": "AKTIF",
                "rapordanTuretildi": True,
            }
            kayitlar.append(kayit)
            kayitDizini[(anahtar, donem)].append(kayit)
            adaylar = [kayit]
            uyar("EKSİK KAYIT", f"{ad} ({donem}): Excel'de kayıt yok, rapordan "
                                f"türetildi (grup {kayit['grupAdi']})")

        raporlar.append({
            "ogrenciAnahtari": anahtar,
            "programAdi": donem,
            "termLabel": donem,
            "groupLabel": sinif,
            "reportDate": pdf_tarihi(yol, yedekTarih),
            "sourcePath": bagil,
            "mutlakYol": str(yol),
            "fileName": yol.name,
            "fileSize": yol.stat().st_size,
        })

    return raporlar


# ---------------------------------------------------------------------------
# Çıktılar
# ---------------------------------------------------------------------------

def raporsuz_listesi(defter: Defter, kayitlar: list[dict],
                     raporlar: list[dict]) -> list[dict]:
    """Kaydı olup o programın raporu bulunmayan öğrenciler."""
    raporluk = {(r["ogrenciAnahtari"], r["programAdi"]) for r in raporlar}
    satirlar = []
    for kayit in kayitlar:
        cift = (kayit["ogrenciAnahtari"], kayit["programAdi"])
        if cift in raporluk:
            continue
        ogr = defter.ogrenciler[kayit["ogrenciAnahtari"]]
        anne = next((v for v in ogr["guardians"] if v["type"] == "ANNE"), None)
        baba = next((v for v in ogr["guardians"] if v["type"] == "BABA"), None)
        satirlar.append({
            "Öğrenci": f"{ogr['firstName']} {ogr['lastName']}",
            "Doğum tarihi": ogr["birthDate"] or "",
            "Okul/Sınıf": ogr["grade"] or "",
            "Program": kayit["programAdi"],
            "Grup": kayit["grupAdi"],
            "Zaman dilimi": "Öğleden Sonra" if kayit["timeSlot"] == "OGLEDEN_SONRA" else "Öğleden Önce",
            "Şube": ogr["branchCode"],
            "Anne": (anne or {}).get("fullName", ""),
            "Anne tel": (anne or {}).get("phone", "") or "",
            "Baba": (baba or {}).get("fullName", ""),
            "Baba tel": (baba or {}).get("phone", "") or "",
        })
    satirlar.sort(key=lambda s: (s["Program"], s["Grup"], s["Öğrenci"]))
    return satirlar


def raporsuz_yaz(satirlar: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raporsuz öğrenciler"
    basliklar = list(satirlar[0].keys()) if satirlar else ["Öğrenci"]
    ws.append(basliklar)
    for hucre in ws[1]:
        hucre.font = Font(bold=True, color="FFFFFF")
        hucre.fill = PatternFill("solid", start_color="4A5568")
        hucre.alignment = Alignment(vertical="center")
    for satir in satirlar:
        ws.append([satir[b] for b in basliklar])
    for sutun in ws.columns:
        en = max(len(str(h.value or "")) for h in sutun)
        ws.column_dimensions[sutun[0].column_letter].width = min(en + 3, 34)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    yol = CIKTI / "raporsuz-ogrenciler.xlsx"
    wb.save(yol)
    return yol


def denetim_yaz(defter: Defter, kayitlar: list[dict], raporlar: list[dict],
                raporsuz: list[dict]) -> Path:
    sayaclar = Counter(u.split("\t", 1)[0] for u in uyarilar)
    programSayisi = Counter(k["programAdi"] for k in kayitlar)
    raporSayisi = Counter(r["programAdi"] for r in raporlar)
    subeSayisi = Counter(o["branchCode"] for o in defter.ogrenciler.values())

    satirlar = [
        "# Geçmiş veri aktarımı — denetim raporu",
        "",
        f"Kaynak Excel: `{EXCEL}`",
        f"Kaynak PDF klasörü: `{PDF_KOK}`",
        "",
        "## Özet",
        "",
        f"- Öğrenci: **{len(defter.ogrenciler)}**",
        f"- Kayıt (öğrenci × program): **{len(kayitlar)}**",
        f"- Arşiv raporu (PDF): **{len(raporlar)}**",
        f"- Toplam PDF boyutu: **{sum(r['fileSize'] for r in raporlar) / 1048576:.0f} MB**",
        f"- Raporu bulunamayan kayıt: **{len(raporsuz)}**",
        "",
        "### Şube dağılımı",
        "",
    ]
    for kod, adet in subeSayisi.most_common():
        satirlar.append(f"- `{kod}`: {adet} öğrenci")

    satirlar += ["", "### Program bazında", "",
                 "| Program | Kayıt | Rapor | Raporsuz |", "| --- | --- | --- | --- |"]
    raporsuzSayisi = Counter(s["Program"] for s in raporsuz)
    for program in sorted(programSayisi):
        satirlar.append(
            f"| {program} | {programSayisi[program]} | {raporSayisi.get(program, 0)} "
            f"| {raporsuzSayisi.get(program, 0)} |"
        )

    satirlar += ["", "## Düzeltme ve uyarı sayıları", ""]
    if sayaclar:
        for kategori, adet in sayaclar.most_common():
            satirlar.append(f"- **{kategori}**: {adet}")
    else:
        satirlar.append("- (yok)")

    satirlar += ["", "## Uygulanan ad eşlemeleri", ""]
    for pdfAdi, excelAdi in sorted(AD_ESLEMESI.items()):
        satirlar.append(f"- `{pdfAdi}` → `{excelAdi}`")

    satirlar += ["", "## Uyarıların tamamı", ""]
    for kategori in sorted(sayaclar):
        satirlar += [f"### {kategori}", ""]
        for uyari in uyarilar:
            if uyari.startswith(kategori + "\t"):
                satirlar.append(f"- {uyari.split(chr(9), 1)[1]}")
        satirlar.append("")

    yol = CIKTI / "denetim.md"
    yol.write_text("\n".join(satirlar) + "\n", encoding="utf-8")
    return yol


def main() -> int:
    if not EXCEL.exists():
        print(f"Excel bulunamadı: {EXCEL}", file=sys.stderr)
        return 1
    if not PDF_KOK.exists():
        print(f"PDF klasörü bulunamadı: {PDF_KOK}", file=sys.stderr)
        return 1
    CIKTI.mkdir(parents=True, exist_ok=True)

    defter = Defter()
    kayitlar = excel_oku(defter)
    raporlar = pdf_oku(defter, kayitlar)
    raporsuz = raporsuz_listesi(defter, kayitlar, raporlar)

    paket = {
        "uretim": {
            "excel": str(EXCEL),
            "pdfKok": str(PDF_KOK),
            "ustBlokSube": UST_BLOK_SUBE,
            "altBlokSube": ALT_BLOK_SUBE,
        },
        "donemler": [
            {"name": ad, "egitimYili": DONEM_EGITIM_YILI[ad]}
            for ad in (D_KIS, D_BAHAR, D_YAZ25, D_YAZ26)
        ],
        "kulupler": [{"name": K_DRAMA}, {"name": K_ROBOTIK}],
        "ogrenciler": list(defter.ogrenciler.values()),
        "kayitlar": kayitlar,
        "raporlar": raporlar,
    }
    veriYolu = CIKTI / "gecmis-veri.json"
    veriYolu.write_text(json.dumps(paket, ensure_ascii=False, indent=1), encoding="utf-8")

    denetimYolu = denetim_yaz(defter, kayitlar, raporlar, raporsuz)
    listeYolu = raporsuz_yaz(raporsuz)

    print(f"Öğrenci        : {len(defter.ogrenciler)}")
    print(f"Kayıt          : {len(kayitlar)}")
    print(f"Arşiv raporu   : {len(raporlar)}")
    print(f"Raporsuz kayıt : {len(raporsuz)}")
    print(f"Uyarı          : {len(uyarilar)}")
    print()
    print(f"  {veriYolu}")
    print(f"  {denetimYolu}")
    print(f"  {listeYolu}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
